import openpyxl
from export import save_to_file
from my_module import *

def converter(FROM: str, TO: str):

    wb = openpyxl.load_workbook(FROM, data_only=True)
    ws = wb.active
    DATE = ws['A2'].value
    try:
        YEAR = DATE.year
    except:
        YEAR = str(DATE).split('-')[0]
    try:
        MONTH = DATE.month
    except:
        MONTH = str(DATE).split('-')[1]

    total_dict = {}

    for rows in list(ws.iter_rows())[1:]:
        if rows[0].value == None:
            continue
        
        # ? 현재 달의 날짜 정보
        day = str(rows[0].value).split('-')[-1]
        if not day.isdigit():
            day = rows[0].value.day
        # -----------------

        # ? 브랜드 이름
        brand = str(rows[1].value)
        brand = brand.strip()
        if not isbrand(brand):
            continue
        # ----------
        
        # ? 발주상품명 (파일 텍스트 그대로)
        order_name = str(rows[2].value)
        order_name = order_name.strip()
        if not isorder(order_name):
            continue

        # ? 메인 이름
        main_name = get_main_name(order_name, brand)

        # ? 서브 정보들
        sub_infos = get_sub_infos(order_name)
        # ? 서브 정보 리스트 (strip 안된)
        sub_info_list = get_sub_info_list(sub_infos)


        # ! 정보 정리하고 dict에 저장
        if sub_info_list == [""]:
            add_to_dict(total_dict, brand, main_name, day, main_name)
        else:
            for sub_info in sub_info_list:
                sub_info = sub_info.strip()
                sub_name = get_sub_name(sub_info)
                mainsub_name = get_mainsub_name(main_name, sub_name)

                add_to_dict(total_dict, brand, mainsub_name, day, sub_info)

    save_to_file(YEAR, MONTH, total_dict, TO)



