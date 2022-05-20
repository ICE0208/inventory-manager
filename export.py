import openpyxl
from openpyxl.styles import Alignment

from my_module import get_days_in_month, num_to_excel_columns

DAY_OFFSET = 2
def save_to_file(year, month, infos: dict, TO: str):
    wb = openpyxl.Workbook() # 임시 엑셀 생성
    sheet = wb.active

    # ? 제목 추가
    title = ['브랜드명', '재고상품명'] + [str(i)+'일' for i in range(1, get_days_in_month(year, month))]\
        + [f"{month}월 총"]
    title_len = len(title)
    sheet.append(title)
    align_center = Alignment(horizontal="center", vertical="center")
    for i in range(1, title_len+1):
        sheet.cell(1, i).alignment = align_center

    # ? 일 추가
    for i, v in enumerate(infos.values()):
        sheet.append(v[:2])
        for day in v[2]:
            sheet.cell(i+2, int(day)+DAY_OFFSET, v[2][day])
            sheet.cell(i+2, int(day)+DAY_OFFSET).alignment = align_center

        #? 월 총
        pos = f"=SUM(C{i+2}:{num_to_excel_columns(title_len-1)}{i+2})"
        sheet.cell(i+2, title_len, pos)
        sheet.cell(i+2, title_len).alignment = align_center

    # ? 가운데 정렬
    for i in range(2, len(infos)+2):
        sheet.cell(i, 1).alignment = align_center
        sheet.cell(i, 2).alignment = align_center
    cur_n = 0
    for day in v[2]:
        cur_n = int(day)+DAY_OFFSET
        cur = num_to_excel_columns(cur_n)
        # ? 일 총
        pos = f"=SUM({cur}2:{cur}{len(infos)+1})"
        sheet.cell(len(infos)+2, cur_n, pos)
        sheet.cell(len(infos)+2, cur_n).alignment = align_center
    cur_n = title_len
    cur = num_to_excel_columns(cur_n)
    pos = f"=SUM({cur}2:{cur}{len(infos)+1})"
    sheet.cell(len(infos)+2, title_len, pos)
    sheet.cell(len(infos)+2, title_len).alignment = align_center

    # ? columns 너비 조정
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 36
    

    wb.save(TO)


